from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def insert(excel_file, sheet_name, predicted_causes):
    """
    Inserts or updates a "Predicted Cause" column next to the "Cause" column in an Excel sheet.

    If a "Predicted Cause" column already exists to the right of the "Cause" column,
    its contents will be overwritten. Otherwise, a new column is inserted.

    The column width is automatically adjusted to fit the content.

    Args:
        excel_file (str): Path to the input Excel (.xlsx) file.
        sheet_name (str): Name of the worksheet to modify.
        predicted_causes (list of str): List of predicted cause values to write, starting from row 2.

    Saves:
        A new Excel file with "_predicted" added before the extension.
    """
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Find the "Cause" column index
    cause_col_idx = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == "Cause":
            cause_col_idx = col[0].column
            break

    if cause_col_idx is None:
        raise ValueError("'Cause' column not found.")

    # Check if the next column already contains "Predicted Cause"
    predicted_col_idx = cause_col_idx + 1
    header_value = ws.cell(row=1, column=predicted_col_idx).value

    if header_value != "Predicted Cause":
        # Insert new column only if needed
        ws.insert_cols(predicted_col_idx)

    # Set header
    ws.cell(row=1, column=predicted_col_idx).value = "Predicted Cause"

    # Fill in the predicted causes
    max_width = len("Predicted Cause")
    for i, value in enumerate(predicted_causes, start=2):
        cell = ws.cell(row=i, column=predicted_col_idx)
        cell.value = value
        if value:
            max_width = max(max_width, len(str(value)))

    # Adjust column width
    col_letter = get_column_letter(predicted_col_idx)
    ws.column_dimensions[col_letter].width = max_width + 2  # Add padding

    # Save to new file
    output_filename = excel_file.replace('.xlsx', '_predicted.xlsx')
    wb.save(output_filename)

    print(f"Cause predictions saved to {output_filename}")
